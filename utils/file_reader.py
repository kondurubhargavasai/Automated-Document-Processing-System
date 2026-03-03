import os
import pdfplumber
from docx import Document
from openpyxl import load_workbook

def read_pdf(file_path):
    text = ""
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            text += page.extract_text() or ""
    return text

def read_docx(file_path):
    doc = Document(file_path)
    return "\n".join([para.text for para in doc.paragraphs])

def read_txt(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        return f.read()

def read_xlsx(file_path):
    wb = load_workbook(file_path)
    text = ""
    for sheet in wb:
        for row in sheet.iter_rows(values_only=True):
            text += " ".join([str(cell) for cell in row if cell]) + "\n"
    return text

def read_file(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == ".pdf":
        return read_pdf(file_path)
    elif ext == ".docx":
        return read_docx(file_path)
    elif ext == ".txt":
        return read_txt(file_path)
    elif ext == ".xlsx":
        return read_xlsx(file_path)
    else:
        return ""